#!/usr/bin/env python3
"""
Generate Excel tables with proper formatting and color-coding for war tracking.
"""

import json
import os
from datetime import datetime
from typing import Dict, List, Set
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WAR_DATA_DIR = "war_data"


def generate_per_war_tables_excel(output_file: str = "per_war_member_performance.xlsx") -> None:
    """
    Generate Excel file with separate tables for each war.
    
    Each war table contains:
    - Member Name
    - Attacks Used
    - Stars Obtained
    - Average Percentage
    - Attack 1 Stars
    - Attack 1 Percentage
    - Attack 2 Stars
    - Attack 2 Percentage
    """
    if not os.path.exists(WAR_DATA_DIR):
        print(f"\n❌ No war data directory found ({WAR_DATA_DIR})")
        return
    
    war_files = sorted([f for f in os.listdir(WAR_DATA_DIR) if f.endswith('.json')])
    
    if not war_files:
        print(f"\n❌ No war data files found in {WAR_DATA_DIR}")
        return
    
    print(f"\n📊 Generating per-war member performance Excel from {len(war_files)} wars...")
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for war_idx, filename in enumerate(war_files, 1):
        filepath = os.path.join(WAR_DATA_DIR, filename)
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Extract war info
            war_start = data.get('war_start', '')
            try:
                war_date = datetime.fromisoformat(war_start.replace('Z', '+00:00')).strftime('%Y-%m-%d')
            except:
                war_date = war_start[:10] if war_start else 'Unknown'
            
            opponent_name = data.get('opponent', {}).get('name', 'Unknown')
            sheet_name = f"War {war_idx} - {opponent_name[:20]}"  # Limit name length
            
            # Create sheet
            ws = wb.create_sheet(sheet_name)
            
            # Add title
            ws.merge_cells('A1:H1')
            title_cell = ws['A1']
            title_cell.value = f"War vs {opponent_name} - {war_date}"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = center_align
            
            # Add headers
            headers = [
                'Member Name',
                'Attacks Used',
                'Stars Obtained',
                'Average %',
                'Attack 1 Stars',
                'Attack 1 %',
                'Attack 2 Stars',
                'Attack 2 %'
            ]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            # Get member data
            member_analysis = data.get('member_analysis', {})
            attackers = member_analysis.get('attackers', [])
            partial = member_analysis.get('partial', [])
            missed = member_analysis.get('missed', [])
            
            # Get raw attack data
            clan_members = data.get('clan', {}).get('members', [])
            
            row_idx = 4
            
            # Process all members (attackers + partial + missed)
            all_members_data = []
            
            # Full attackers
            for member in attackers:
                name = member['name']
                attacks_used = member['attacks_used']
                total_stars = member['total_stars']
                avg_percent = member['avg_destruction_percent']
                
                # Get individual attacks
                member_attacks = None
                for m in clan_members:
                    if m.get('name') == name:
                        member_attacks = m.get('attacks', [])
                        break
                
                attack1_stars = ''
                attack1_percent = ''
                attack2_stars = ''
                attack2_percent = ''
                
                if member_attacks:
                    if len(member_attacks) >= 1:
                        attack1_stars = member_attacks[0].get('stars', 0)
                        attack1_percent = f"{member_attacks[0].get('destructionPercentage', 0):.1f}"
                    if len(member_attacks) >= 2:
                        attack2_stars = member_attacks[1].get('stars', 0)
                        attack2_percent = f"{member_attacks[1].get('destructionPercentage', 0):.1f}"
                
                all_members_data.append({
                    'name': name,
                    'attacks_used': attacks_used,
                    'total_stars': total_stars,
                    'avg_percent': f"{avg_percent:.1f}",
                    'attack1_stars': attack1_stars,
                    'attack1_percent': attack1_percent,
                    'attack2_stars': attack2_stars,
                    'attack2_percent': attack2_percent
                })
            
            # Partial attackers
            for member in partial:
                name = member['name']
                attacks_used = member['attacks_used']
                total_stars = member['total_stars']
                avg_percent = member['avg_destruction_percent']
                
                # Get individual attacks
                member_attacks = None
                for m in clan_members:
                    if m.get('name') == name:
                        member_attacks = m.get('attacks', [])
                        break
                
                attack1_stars = ''
                attack1_percent = ''
                attack2_stars = ''
                attack2_percent = ''
                
                if member_attacks:
                    if len(member_attacks) >= 1:
                        attack1_stars = member_attacks[0].get('stars', 0)
                        attack1_percent = f"{member_attacks[0].get('destructionPercentage', 0):.1f}"
                    if len(member_attacks) >= 2:
                        attack2_stars = member_attacks[1].get('stars', 0)
                        attack2_percent = f"{member_attacks[1].get('destructionPercentage', 0):.1f}"
                
                all_members_data.append({
                    'name': name,
                    'attacks_used': attacks_used,
                    'total_stars': total_stars,
                    'avg_percent': f"{avg_percent:.1f}",
                    'attack1_stars': attack1_stars,
                    'attack1_percent': attack1_percent,
                    'attack2_stars': attack2_stars,
                    'attack2_percent': attack2_percent
                })
            
            # Missed attacks
            for member in missed:
                all_members_data.append({
                    'name': member['name'],
                    'attacks_used': 0,
                    'total_stars': 0,
                    'avg_percent': '0.0',
                    'attack1_stars': '',
                    'attack1_percent': '',
                    'attack2_stars': '',
                    'attack2_percent': ''
                })
            
            # Sort by total stars (descending), then by name
            all_members_data.sort(key=lambda x: (-x['total_stars'], x['name']))
            
            # Write data
            for member_data in all_members_data:
                ws.cell(row=row_idx, column=1).value = member_data['name']
                ws.cell(row=row_idx, column=2).value = member_data['attacks_used']
                ws.cell(row=row_idx, column=3).value = member_data['total_stars']
                ws.cell(row=row_idx, column=4).value = member_data['avg_percent']
                ws.cell(row=row_idx, column=5).value = member_data['attack1_stars']
                ws.cell(row=row_idx, column=6).value = member_data['attack1_percent']
                ws.cell(row=row_idx, column=7).value = member_data['attack2_stars']
                ws.cell(row=row_idx, column=8).value = member_data['attack2_percent']
                
                # Apply formatting
                for col in range(1, 9):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = thin_border
                    if col >= 2:  # Align numbers to center
                        cell.alignment = center_align
                
                row_idx += 1
            
            # Auto-size columns
            for col_idx in range(1, 9):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in ws[col_letter]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2
            
        except Exception as e:
            print(f"⚠️ Error processing {filename}: {e}")
            continue
    
    # Save workbook
    wb.save(output_file)
    print(f"✅ Per-war member performance Excel saved to: {output_file}")
    print(f"   Total wars: {len(wb.sheetnames)} sheets")


def generate_overall_member_performance_excel(output_file: str = "overall_member_performance.xlsx") -> None:
    """
    Generate Excel file with overall member performance across all wars.
    
    Columns:
    - Member Name
    - Wars Participated
    - Wars Missed (attacks)
    - Average Stars per War
    - Average Percentage per War
    
    Color coding:
    - Yellow: 2 consecutive missed wars
    - Red: 3+ consecutive missed wars
    """
    if not os.path.exists(WAR_DATA_DIR):
        print(f"\n❌ No war data directory found ({WAR_DATA_DIR})")
        return
    
    war_files = sorted([f for f in os.listdir(WAR_DATA_DIR) if f.endswith('.json')])
    
    if not war_files:
        print(f"\n❌ No war data files found in {WAR_DATA_DIR}")
        return
    
    print(f"\n📊 Generating overall member performance Excel from {len(war_files)} wars...")
    
    # Track member stats: {member_name: {...}}
    member_stats = {}
    # Track consecutive misses: {member_name: [list of bools (True=participated)]}
    member_war_history = {}
    
    for filename in war_files:
        filepath = os.path.join(WAR_DATA_DIR, filename)
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            member_analysis = data.get('member_analysis', {})
            attackers = member_analysis.get('attackers', [])
            partial = member_analysis.get('partial', [])
            missed = member_analysis.get('missed', [])
            
            # Get all members in this war
            all_war_members = set()
            
            # Process attackers (full participation)
            for member in attackers:
                name = member['name']
                all_war_members.add(name)
                
                if name not in member_stats:
                    member_stats[name] = {
                        'wars_participated': 0,
                        'wars_missed_attacks': 0,
                        'total_stars': 0,
                        'total_destruction': 0.0,
                        'total_attacks': 0
                    }
                    member_war_history[name] = []
                
                member_stats[name]['wars_participated'] += 1
                member_stats[name]['total_stars'] += member['total_stars']
                member_stats[name]['total_destruction'] += member['avg_destruction_percent'] * member['attacks_used']
                member_stats[name]['total_attacks'] += member['attacks_used']
                member_war_history[name].append(True)  # Participated fully
            
            # Process partial attackers
            for member in partial:
                name = member['name']
                all_war_members.add(name)
                
                if name not in member_stats:
                    member_stats[name] = {
                        'wars_participated': 0,
                        'wars_missed_attacks': 0,
                        'total_stars': 0,
                        'total_destruction': 0.0,
                        'total_attacks': 0
                    }
                    member_war_history[name] = []
                
                member_stats[name]['wars_participated'] += 1
                member_stats[name]['wars_missed_attacks'] += 1  # Partial = missed some attacks
                member_stats[name]['total_stars'] += member['total_stars']
                member_stats[name]['total_destruction'] += member['avg_destruction_percent'] * member['attacks_used']
                member_stats[name]['total_attacks'] += member['attacks_used']
                member_war_history[name].append(False)  # Didn't use all attacks
            
            # Process missed attacks (no attacks)
            for member in missed:
                name = member['name']
                all_war_members.add(name)
                
                if name not in member_stats:
                    member_stats[name] = {
                        'wars_participated': 0,
                        'wars_missed_attacks': 0,
                        'total_stars': 0,
                        'total_destruction': 0.0,
                        'total_attacks': 0
                    }
                    member_war_history[name] = []
                
                member_stats[name]['wars_participated'] += 1
                member_stats[name]['wars_missed_attacks'] += 1
                member_war_history[name].append(False)  # Missed attacks
            
        except Exception as e:
            print(f"⚠️ Error processing {filename}: {e}")
            continue
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Overall Performance"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "Overall Member Performance Across All Wars"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_align
    
    # Add headers
    headers = [
        'Member Name',
        'Wars Participated',
        'Wars Missed (Attacks)',
        'Average Stars per War',
        'Average % per War'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Prepare data rows
    rows_data = []
    for name, stats in member_stats.items():
        wars_participated = stats['wars_participated']
        wars_missed = stats['wars_missed_attacks']
        total_stars = stats['total_stars']
        total_destruction = stats['total_destruction']
        total_attacks = stats['total_attacks']
        
        # Calculate averages per war
        avg_stars_per_war = (total_stars / wars_participated) if wars_participated > 0 else 0
        avg_percent_per_war = (total_destruction / total_attacks) if total_attacks > 0 else 0
        
        # Check for consecutive misses
        consecutive_misses = 0
        max_consecutive = 0
        for participated in reversed(member_war_history[name]):
            if not participated:
                consecutive_misses += 1
                max_consecutive = max(max_consecutive, consecutive_misses)
            else:
                break  # Only count most recent consecutive
        
        rows_data.append({
            'name': name,
            'wars_participated': wars_participated,
            'wars_missed': wars_missed,
            'avg_stars': f"{avg_stars_per_war:.2f}",
            'avg_percent': f"{avg_percent_per_war:.1f}",
            'consecutive_misses': consecutive_misses
        })
    
    # Sort by wars participated (descending), then by name
    rows_data.sort(key=lambda x: (-x['wars_participated'], x['name']))
    
    # Write data with color coding
    row_idx = 4
    for row_data in rows_data:
        ws.cell(row=row_idx, column=1).value = row_data['name']
        ws.cell(row=row_idx, column=2).value = row_data['wars_participated']
        ws.cell(row=row_idx, column=3).value = row_data['wars_missed']
        ws.cell(row=row_idx, column=4).value = row_data['avg_stars']
        ws.cell(row=row_idx, column=5).value = row_data['avg_percent']
        
        # Apply color coding based on consecutive misses
        fill = None
        if row_data['consecutive_misses'] >= 3:
            fill = red_fill
        elif row_data['consecutive_misses'] == 2:
            fill = yellow_fill
        
        # Apply formatting
        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            if col >= 2:  # Align numbers to center
                cell.alignment = center_align
            if fill and col == 1:  # Apply color to name column
                cell.fill = fill
        
        row_idx += 1
    
    # Auto-size columns
    for col_idx in range(1, 6):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2
    
    # Add legend
    legend_row = row_idx + 2
    ws.cell(row=legend_row, column=1).value = "Legend:"
    ws.cell(row=legend_row, column=1).font = Font(bold=True)
    
    ws.cell(row=legend_row + 1, column=1).value = "Yellow"
    ws.cell(row=legend_row + 1, column=1).fill = yellow_fill
    ws.cell(row=legend_row + 1, column=2).value = "2 consecutive missed wars"
    
    ws.cell(row=legend_row + 2, column=1).value = "Red"
    ws.cell(row=legend_row + 2, column=1).fill = red_fill
    ws.cell(row=legend_row + 2, column=2).value = "3+ consecutive missed wars"
    
    # Save workbook
    wb.save(output_file)
    print(f"✅ Overall member performance Excel saved to: {output_file}")
    print(f"   Total members: {len(rows_data)}")


def generate_all_excel_tables() -> None:
    """Generate both Excel tables."""
    print("\n" + "=" * 80)
    print("📊 GENERATING ALL WAR ANALYSIS EXCEL FILES")
    print("=" * 80)
    
    generate_per_war_tables_excel()
    generate_overall_member_performance_excel()
    
    print("\n✅ All Excel files generated successfully!")
    print("\nGenerated files:")
    print("  1. per_war_member_performance.xlsx (separate sheet per war)")
    print("  2. overall_member_performance.xlsx (with color-coding)")


if __name__ == "__main__":
    generate_all_excel_tables()
